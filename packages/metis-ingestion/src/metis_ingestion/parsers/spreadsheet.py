"""Spreadsheets: xlsx via openpyxl, csv decoded as-is. Rows become TSV lines."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from metis_ingestion._text import decode_text, normalize_blocks


def extract_xlsx(data: bytes) -> str:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheets: list[str] = []
        for worksheet in workbook.worksheets:
            lines: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = ["" if value is None else str(value) for value in row]
                if any(cell for cell in cells):
                    lines.append("\t".join(cells))
            if lines:
                sheets.append(f"# {worksheet.title}\n" + "\n".join(lines))
        return normalize_blocks("\n\n".join(sheets))
    finally:
        workbook.close()


def extract_csv(data: bytes) -> str:
    return normalize_blocks(decode_text(data))
